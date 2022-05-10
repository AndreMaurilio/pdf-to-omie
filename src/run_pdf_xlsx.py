# Script que extrai dados do pdf e preenche a planilha do padrao Omie de lançamento de notas.


import os

import openpyxl
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage

WB = ''
PDF_PATH = ''
CNPJ = ''
NPC = ''
CODEMICRO = {}



def write_pdf_txt(path, pdf):
    with open(path, mode="w") as output_file:
        for page in pdf.pages:
            text = page.extractText()
            output_file.write(text)


def pdf_to_xlsx(wb, pdf, address):
    sh1 = wb.active
    row = sh1.max_row
    r = 22
    global CNPJ
    #   print(wb['Omie_Pedido_Venda'].cell(3, 17).value)
    sh1.cell(row=7, column=4, value=CNPJ)
    sh1.cell(row=7, column=6, value="Suprimentos")
    sh1.cell(row=7, column=7, value="Para 21 dias ")
    sh1.cell(row=7, column=10, value="Banco Bradesco")


    CNPJ = ''

    for i in range(r, row + 1):
        for p in range(len(pdf)):
            for k in range(len(pdf[p])):
                pedido = pdf[p][k]
                sh1.cell(row=i, column=3, value=pedido["Material"])
                sh1.cell(row=i, column=4, value=pedido["Produto"])
                sh1.cell(row=i, column=5, value="MAXION (À FATURAR)")
                sh1.cell(row=i, column=6, value=pedido["QTD"])
                sh1.cell(row=i, column=7, value=pedido["Preço"])
                sh1.cell(row=i, column=10, value=pedido["Pedido de Compras"])
                sh1.cell(row=i, column=11, value=pedido["item"])

                i += 1
            p += 1
        else:
            break
    wb.save(os.getcwd() + "\\xlsx\\" + address[:-4] + ".xlsx")


def represent_int(s):
    n = s
    try:
        int(n)
        return True
    except ValueError:
        return False


def get_pdf_miner_file(path, wb):
    resource_manager = PDFResourceManager(caching=True)
    out_text = StringIO()
    laParams = LAParams()
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(path, 'rb')
    ar = list()
    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password='', caching=True, check_extractable=True):
        interpreter.process_page(page)
    text = out_text.getvalue()
    text = text.split("\n")
    ar.append(collect_pdf_data(text))
    address = path.split("/")[-1]
    pdf_to_xlsx(wb, ar, address)


def collect_pdf_data(lista):
    pedidos = list()
    precos = list()
    n = 0
    global CNPJ
    global NPC
    for i in lista:
        if len(i) > 9:
            if i[:9] == 'C.N.P.J.:' and CNPJ == '':
                CNPJ = i[10:]
        if len(i) == 10 and represent_int(i) and NPC == '':
            NPC = i
        if len(i) == 17:
            i = i.split(" ")
            if len(i) == 2:
                if represent_int(i[0]) and represent_int(i[1]):
                    pedidos.append(create_item(i[0], i[1], NPC, '', '', False))

        if ("PEÇ" in i or "UN" in i) and (8 <= len(i) <= 10):
            i = i.split(" ")
            precos.append(get_price_and_qtd(i[0], lista[n + 2]))
        n += 1
    pedidos = correct_price_and_qtd(pedidos, precos)
    NPC = ''
    return pedidos


def correct_price_and_qtd(lista, incomplete):
    for i in lista:
        if not i["complete"] and len(incomplete) > 0:
            i["QTD"] = incomplete[0]["QTD"]
            i["Preço"] = incomplete[0]["Preço"]
            incomplete.pop(0)
    return lista


def load_codes_item():
    with open(os.getcwd() + "/codigos.txt", "r") as file_in:
        for line in file_in:
            CODEMICRO[line.split(":")[0]] = line.split(":")[1][:-1]

def create_item(item, material, pedidocompra, qtd, preco, complete):
    produto = ""
    if material in CODEMICRO:
        produto = CODEMICRO[material]

    return {"item": item, "Material": material, "Produto": produto, "Pedido de Compras": pedidocompra, "QTD": qtd,
            "Frete": "", "Preço": preco, "complete": complete}


def get_price_and_qtd(qtd, preco):
    return {"QTD": qtd, "Preço": preco};


def pdf_to_omie_xlsx(path, pathomie, label):
    global WB, PDF_PATH
    PDF_PATH = path
    load_codes_item()
    WB = openpyxl.load_workbook(pathomie)
    get_pdf_miner_file(PDF_PATH, WB)
    return "Fim!!"
