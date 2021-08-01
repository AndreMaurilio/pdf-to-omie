# Script que extrai dados do pdf e preenche a planilha do padrao Omie de lançamento de notas.


import io

import openpyxl
from PyPDF2 import PdfFileReader
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage

WB = openpyxl.load_workbook("/home/andre/Documentos/workspace/Python-Projects/pdf-to-omie/Omie_EXTRAIDO.xlsx")
PDF_PATH = '/home/andre/Documentos/workspace/Python-Projects/pdf-to-omie/MAXION.PDF'  # MAXION.PDF, 4500022260.pdf
TXT_PATH = '/home/andre/Downloads/pdf_extraido.txt'
CNPJ = ''
NPC = ''


def read_pdf(path):
    # Le o arquivo
    pdf = PdfFileReader(str(path))
    # Printa o numero de paginas
    print(pdf.getNumPages())
    # printa todos os textos
    # for page in pdf.pages:
    #     print(page.extractText(), end='\n')
    return pdf


def write_pdf_txt(path, pdf):
    with open(path, mode="w") as output_file:
        for page in pdf.pages:
            text = page.extractText()
            output_file.write(text)


def pdf_to_xlsx(wb, pdf):
    sh1 = wb.active
    row = sh1.max_row
    r = 22
    global CNPJ
    #   print(wb['Omie_Pedido_Venda'].cell(3, 17).value)
    sh1.cell(row=7, column=4, value=CNPJ)
    CNPJ = ''
    for i in range(r, row + 1):
        for p in range(len(pdf)):
            for k in range(len(pdf[p])):
                pedido = pdf[p][k]
                sh1.cell(row=i, column=11, value=pedido["item"])
                sh1.cell(row=i, column=3, value=pedido["Material"])
                sh1.cell(row=i, column=10, value=pedido["Pedido de Compras"])
                sh1.cell(row=i, column=6, value=pedido["QTD"])
                # sh1.cell(row=i, column=20, value=pedido["Frete"])
                sh1.cell(row=i, column=7, value=pedido["Preço"])
                i += 1
            p += 1
        else:
            break
    wb.save("/home/andre/Documentos/workspace/Python-Projects/pdf-to-omie/Omie_EXTRAIDO.xlsx")


def print_pdf_txt(path, pdf, wb, cnpj):
    with open(path, mode="w"):
        ar = list()
        for page in pdf.pages:
            text = page.extractText()
            text = text.split(
                "ItemDescrição MaterialPedido deCompras /Programa deremessaQUANT.UMPreçoUnitárioFreteUtilização do MaterialMoedaPreço TotalSaldoData de Entrega")
            if (isinstance(text, list) and len(text) > 1):
                ar.append(populate_pedidos(text[1]))
            else:
                print("DESCARTADO \n")
    pdf_to_xlsx(wb, ar, cnpj)


def split_item_desc(st):
    n = 16
    while (n + 16) < len(st):
        if represent_int(st[n:n + 16]):
            st = list(st)
            st[n - 1] = '&'
            st = ''.join(st)
        n += 1;
    v = st.split("&")
    return v


def populate_pedidos(lista):
    lista = split_item_desc(lista)
    pedidos = list()
    if isinstance(lista, list):
        for s in lista:
            aux = s[0:16]
            if (represent_int(aux) == True):
                item = s[0:5]
                material = s[5:16]
                st = ""
                n2 = 17
                stop = False
                for c in s[16:]:
                    n = 0
                    st = c
                    for e in s[n2:]:

                        if n == 9:
                            if represent_int(st):
                                pedidocompra = st
                                stop = True
                                break
                            else:
                                n = 0
                                st = ""
                                break
                        st = st + e
                        n += 1

                    if stop:
                        break
                    n2 = n2 + 1

                if "PEÇ" in s[n2:n2 + 30]:
                    v = s[n2:].split("PEÇ")
                else:
                    v = s[n2:].split("UN")
                qtd = v[0][9:]
                if len(v) > 1:
                    preco = v[1][:6]
                else:
                    preco = "nd"
                pedidos.append({"item": item, "Material": material, "Pedido de Compras": pedidocompra,
                                "QTD": qtd, "Frete": "", "Preço": preco})

    return pedidos


def represent_int(s):
    n = s
    try:
        int(n)
        return True
    except ValueError:
        return False


def get_pdf_file(path, exc, wb):
    resource_manager = PDFResourceManager(caching=True)
    out_text = StringIO()
    laParams = LAParams()
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password='', caching=True, check_extractable=True):
        interpreter.process_page(page)
    text = out_text.getvalue()
    fp.close()
    text_converter.close()
    out_text.close()
    return text


def get_pdf_miner_file(path, wb):
    resource_manager = PDFResourceManager(caching=True)
    out_text = StringIO()
    laParams = LAParams()
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(path, 'rb')
    ar = list()
    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password='', caching=True, check_extractable=True):
        interpreter.process_page(page)
    text = out_text.getvalue()
    text = text.split("\n")
    ar.append(collect_pdf_data(text))
    pdf_to_xlsx(wb, ar)


def collect_pdf_data(lista):
    pedidos = list()
    precos = list()
    n = 0
    global CNPJ
    global NPC
    for i in lista:
        if len(i) > 9:
            if i[:9] == 'C.N.P.J.:' and CNPJ == '':
                CNPJ = i[10:]
        if len(i) == 10 and represent_int(i) and NPC == '':
            NPC = i
        if len(i) == 17:
            i = i.split(" ")
            if len(i) == 2:
                if represent_int(i[0]) and represent_int(i[1]):
                    pedidos.append(create_item(i[0], i[1], NPC, '', '', False))

        if ("PEÇ" in i or "UN" in i) and (8 <= len(i) <= 10):
            i = i.split(" ")
            precos.append(get_price_and_qtd(i[0], lista[n + 2]))
        n += 1
    pedidos = correct_price_and_qtd(pedidos, precos)
    return pedidos


def correct_price_and_qtd(lista, incomplete):
    for i in lista:
        if not i["complete"] and len(incomplete) > 0:
            i["QTD"] = incomplete[0]["QTD"]
            i["Preço"] = incomplete[0]["Preço"]
            incomplete.pop(0)
    return lista


def create_item(item, material, pedidocompra, qtd, preco, complete):
    return {"item": item, "Material": material, "Pedido de Compras": pedidocompra, "QTD": qtd,
            "Frete": "", "Preço": preco, "complete": complete};


def get_price_and_qtd(qtd, preco):
    return {"QTD": qtd, "Preço": preco};


def is_complete(item, material, pedidocompra, qtd, preco):
    if item != '' and material != '' and qtd != '' and preco != '' and pedidocompra != '':
        return True
    else:
        return False


if __name__ == '__main__':
    get_pdf_miner_file(PDF_PATH, WB)
    print("FIM!!")
